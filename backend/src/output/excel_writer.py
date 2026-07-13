import json
import os
import shutil
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

class ExcelWriter:
    def __init__(self, mapping_path: str):
        with open(mapping_path, 'r') as f:
            self.mapping = json.load(f)

    def write_consolidated(self, data_list: list, template_path: str, output_path: str):
        try:
            wb = openpyxl.load_workbook(template_path)
        except FileNotFoundError:
            raise FileNotFoundError(f"Template {template_path} not found.")

        if "Current Plans_Template" not in wb.sheetnames:
            template_sheet = wb.active
            template_sheet.title = "Current Plans_Template"
        else:
            template_sheet = wb["Current Plans_Template"]

        for data in data_list:
            carrier = data.get('plan_information', {}).get('carrier', 'Carrier')
            plan_name = data.get('plan_information', {}).get('plan_name', 'Plan')
            
            safe_name = f"{carrier} {plan_name}"
            safe_name = "".join(c for c in safe_name if c not in r"\/?*[]:")[:31]

            new_sheet = wb.copy_worksheet(template_sheet)
            new_sheet.title = safe_name
            
            print(f"      ↳ Creating sheet: '{safe_name}'")

            # Set up data validation rules for modifiers and checkbox-like fields
            dv_medical = DataValidation(type="list", formula1='"After Deductible,Deductible Waived"', allow_blank=True)
            dv_pharmacy = DataValidation(type="list", formula1='"Rx - After Plan Deductible,Rx - After Rx Deductible,Rx - Deductible Waived"', allow_blank=True)
            dv_checkbox = DataValidation(type="list", formula1='"☑,☐"', allow_blank=True)

            # Define specific dropdown fields to ensure they have validations
            dropdown_fields = {
                "plan_information.carrier",
                "plan_information.plan_source",
                "plan_information.plan_type",
                "plan_information.specialty_network",
                "deductibles_and_coinsurance.deductible_type",
                "deductibles_and_coinsurance.family_tier",
            }
            checkbox_fields = {
                "plan_information.hdhp",
                "plan_information.open_access",
                "pharmacy.offers_tier_1a_benefit",
                "pharmacy.specialty_mirrors_tiers_1_3",
                "out_of_network.out_of_network_coverage",
            }

            cells_populated = 0
            for field_path, cell_ref in self.mapping.items():
                keys = field_path.split('.')
                val = data
                for k in keys:
                    if isinstance(val, dict):
                        val = val.get(k)
                    else:
                        val = getattr(val, k, None)

                if field_path in checkbox_fields:
                    if isinstance(val, bool):
                        val = "☑" if val else "☐"
                    elif isinstance(val, str):
                        normalized = val.strip().lower()
                        if normalized in {"true", "yes", "y", "1", "checked", "☑"}:
                            val = "☑"
                        elif normalized in {"false", "no", "n", "0", "unchecked", "☐"}:
                            val = "☐"

                if val is None or val == "null" or str(val).strip() == "":
                    if "modifier" not in field_path and "status" not in field_path and "type" not in field_path and field_path not in dropdown_fields:
                        if "coinsurance" in field_path:
                            val = "0%"
                        elif "copay" in field_path or "deductible" in field_path or "oop_max" in field_path:
                            val = "$0"
                        else:
                            val = ""
                    else:
                        # For status/modifier/type fields that are null, skip writing entirely
                        val = None

                if val is not None and val != "null" and str(val).strip() != "":
                    # Status fields (Column H): write as plain text, no numeric stripping
                    if "status" in field_path:
                        new_sheet[cell_ref] = str(val).strip()
                        cells_populated += 1

                    # For all other fields, attempt numeric conversion only if it looks like a pure number
                    # DO NOT convert if the value contains currency symbols or text
                    elif field_path not in checkbox_fields and "modifier" not in field_path and "status" not in field_path and "type" not in field_path and "name" not in field_path and "notes" not in field_path and "description" not in field_path and "date" not in field_path and "source" not in field_path and "carrier" not in field_path and "network" not in field_path and field_path != "plan_information.zip_code" and field_path != "plan_information.group_number" and field_path not in dropdown_fields:
                        sval = str(val).strip()
                        
                        # CRITICAL: If value starts with $ or contains %, write as-is (formatted currency/percentage)
                        # These should NOT be converted to numeric values
                        if sval.startswith('$') or sval.endswith('%') or '/' in sval or ' or ' in sval.lower():
                            # Keep as text - these are formatted currency, percentages, or complex expressions
                            new_sheet[cell_ref] = sval
                            cells_populated += 1
                        elif any(c.isdigit() for c in sval):
                            # Only try numeric conversion for pure numeric-looking strings
                            import re
                            cleaned = re.sub(r'[^\d.]', '', sval)
                            if cleaned:
                                try:
                                    num = float(cleaned)
                                    # Use int if it's a whole number
                                    if num.is_integer():
                                        val = int(num)
                                    else:
                                        val = num
                                except ValueError:
                                    val = sval
                            new_sheet[cell_ref] = val
                            cells_populated += 1
                        else:
                            new_sheet[cell_ref] = sval
                            cells_populated += 1
                    else:
                        new_sheet[cell_ref] = val
                        cells_populated += 1
                elif val == "":
                    new_sheet[cell_ref] = ""

                # Apply data validation dropdowns to modifier cells and checkbox fields
                if field_path.endswith("_modifier"):
                    if field_path.startswith("pharmacy."):
                        dv_pharmacy.add(new_sheet[cell_ref])
                    else:
                        dv_medical.add(new_sheet[cell_ref])
                elif field_path in checkbox_fields:
                    dv_checkbox.add(new_sheet[cell_ref])
                        
            # Add validations to the sheet after all cells are registered
            new_sheet.add_data_validation(dv_medical)
            new_sheet.add_data_validation(dv_pharmacy)
            new_sheet.add_data_validation(dv_checkbox)
            
            # Recreate the dropdown field validations since wb.copy_worksheet drops them
            dv_carrier = DataValidation(type="list", formula1="Lists!$A$2:$A$64", allow_blank=True)
            dv_carrier.add(new_sheet["D14"])
            new_sheet.add_data_validation(dv_carrier)
            
            dv_plan_source = DataValidation(type="list", formula1="Lists!$B$2:$B$130", allow_blank=True)
            dv_plan_source.add(new_sheet["D15"])
            new_sheet.add_data_validation(dv_plan_source)
            
            dv_plan_type = DataValidation(type="list", formula1="Lists!$C$2:$C$4", allow_blank=True)
            dv_plan_type.add(new_sheet["D16"])
            new_sheet.add_data_validation(dv_plan_type)
            
            dv_specialty_network = DataValidation(type="list", formula1="Lists!$D$2:$D$7", allow_blank=True)
            dv_specialty_network.add(new_sheet["D17"])
            new_sheet.add_data_validation(dv_specialty_network)
            
            dv_deductible_type = DataValidation(type="list", formula1="Lists!$E$2:$E$4", allow_blank=True)
            dv_deductible_type.add(new_sheet["D31"])
            new_sheet.add_data_validation(dv_deductible_type)
            
            dv_family_tier = DataValidation(type="list", formula1="Lists!$F$2:$F$6", allow_blank=True)
            dv_family_tier.add(new_sheet["D32"])
            new_sheet.add_data_validation(dv_family_tier)
            
            dv_formulary_type = DataValidation(type="list", formula1="Lists!$I$2:$I$5", allow_blank=True)
            dv_formulary_type.add(new_sheet["D85"])
            new_sheet.add_data_validation(dv_formulary_type)
            
            print(f"      ↳ Cells mapped:   {cells_populated}/{len(self.mapping)}")

        # Remove the blank template sheet so only populated sheets remain
        wb.remove(template_sheet)
        
        # Hide the 'Lists' sheet so it doesn't clutter the output file
        # (We cannot delete it, otherwise the dropdowns will break)
        if "Lists" in wb.sheetnames:
            wb["Lists"].sheet_state = "hidden"

        # Write to temp file first, then move (avoids lock issues on Windows)
        temp_path = output_path + ".tmp"
        wb.save(temp_path)
        
        # Replace old file with temp (shutil.move handles locked files better)
        try:
            if os.path.exists(output_path):
                try:
                    os.remove(output_path)
                except PermissionError:
                    # If locked, just overwrite
                    pass
            shutil.move(temp_path, output_path, copy_function=shutil.copy2)
        except Exception as e:
            # Fallback: just save as temp was successful, leave it
            if os.path.exists(temp_path):
                try:
                    shutil.move(temp_path, output_path, copy_function=shutil.copy2)
                except:
                    pass  # Temp file saved successfully, that's enough
